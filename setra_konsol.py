"""
SETRA Bölüm 5 — Dengeli Konsol Stabilite Hesabı
Tkinter GUI  ·  PDF + Excel çıktısı
BPEL 91 / BAEL 91 / SETRA GB Chapter 5

Çalıştırma:
    python setra_konsol.py

Gerekli paketler:
    pip install reportlab openpyxl
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import math
import os
import datetime


# ═══════════════════════════════════════════════════════════════════════════════
#  HESAP MOTORU
# ═══════════════════════════════════════════════════════════════════════════════

class SetraHesap:
    """
    SETRA §5 — Dengeli Konsol Stabilite Hesap Motoru

    §5.3.1  Krawsky formülleri → ağırlık ve ağırlık merkezi
    §5.3.2  Değişken inşaat yükleri (Q_PRC, Q_PRA, Q_w)
    §5.3.3  Kazara yük F_A
    §5.4.1  Tip A kombinasyonları (Geçici inşaat)
    §5.4.2  Tip B kombinasyonları (Kazara)
    §5.5.1  Tendon sayısı ve kuvveti
    §5.5.2  Blok yüzey alanı
    §5.5.3  Ayak dönme açısı
    """

    def __init__(self, p: dict):
        self.p = p
        self.sonuclar: dict = {}

    def hesapla(self) -> dict:
        p = self.p
        s: dict = {}

        lf        = p["lf"]
        B1        = p["B1"]
        B0        = p["B0"]
        gam       = p["gamma"]
        bslab     = p["bslab"]
        e         = p["e_blk"]
        d         = p["d_blk"]
        ploss     = p["p_loss"] / 100.0
        fprg      = p["fprg"]
        fpeg      = p["fpeg"]
        sten      = p["s_ten"] / 1e6      # mm² → m²
        hp        = p["hp"]
        Ip        = p["Ip"]
        fc28      = p["fc28"]
        ba        = p["blk_a"]
        bb        = p["blk_b"]
        a0        = p["a0"]
        b0        = p["b0"]
        Qprc1     = p["Qprc1"]
        dex       = p["dex"]
        Qpra1_kPa = p["Qpra1"]
        Qpra2_usr = p["Qpra2"]
        Qw_kPa    = p["Qw"]
        L_ten     = p["L_ten"]
        Es_GPa    = p["Es"]

        # ── §5.3.1 Krawsky ────────────────────────────────────────────────
        P   = (B1 + 2*B0) * gam * lf / 3.0
        dg  = (B1 + 5*B0) * lf / (4*(B1 + 2*B0))
        G_max = 1.02 * P
        G_min = 0.98 * P

        # ── §5.3.2.1 ──────────────────────────────────────────────────────
        Qprc1max = 1.06 * Qprc1
        Qprc1min = 0.96 * Qprc1

        # ── §5.3.2.2 ──────────────────────────────────────────────────────
        Qpra1_kN = Qpra1_kPa * bslab * lf
        Qpra2    = Qpra2_usr if Qpra2_usr > 0 else (50 + 5*bslab)
        Qw_kN    = Qw_kPa * bslab * lf

        # ── §5.3.3 Kazara ─────────────────────────────────────────────────
        FA = 2.0 * Qprc1max   # dinamik katsayı ×2

        # MN birimine çevir
        def mn(x): return x / 1000.0
        gmax = mn(G_max);  gmin = mn(G_min)
        qmax = mn(Qprc1max); qmin = mn(Qprc1min)
        qra1 = mn(Qpra1_kN); qra2 = mn(Qpra2); qw = mn(Qw_kN)
        fa   = mn(FA)

        # ── §5.4.1 Tip A Kombinasyonları ──────────────────────────────────
        N_A1 = 1.1 * (gmax + gmin)
        N_A2 = 0.9 * (gmax + gmin)
        M_G  = (gmax - gmin) * dg
        M_var_A = 1.25 * ((qmax - qmin)*lf
                          + qra1 * lf/2
                          + qra2 * (lf - dex)
                          + qw  * lf/2)
        M_A1 = abs(1.1 * M_G + M_var_A)
        M_A2 = abs(0.9 * M_G + M_var_A)

        # ── §5.4.2 Tip B Kombinasyonları ──────────────────────────────────
        N_B1 = 1.1 * (gmax + gmin)
        N_B2 = 0.9 * (gmax + gmin)
        M_var_B = (fa * lf
                   + 1.0 * (qmax * lf
                             + qra1 * lf/2
                             + qra2 * (lf - dex)))
        M_B1 = abs(1.1 * M_G + M_var_B)
        M_B2 = abs(0.9 * M_G + M_var_B)

        # ── §5.5.1 σ_p0 ───────────────────────────────────────────────────
        sp0 = min(0.8*fprg, 0.9*fpeg) / 1000.0   # MN/m²

        # ── §5.5.1.1 A kombinasyonu → tendon boyutlandırma ────────────────
        # n×s = max(0, M/e − N/2) / [(1−p)·σ_p0]
        def nxs_A_f(M, N):
            return max(0.0, M/e - N/2) / ((1 - ploss) * sp0)

        nxs_A_req = max(nxs_A_f(M_A1, N_A1), nxs_A_f(M_A2, N_A2))

        n_raw = nxs_A_req / sten if sten > 0 else 0
        n_req = max(2, math.ceil(n_raw / 2) * 2)   # çift sayı, min 2

        Fi  = n_req * (1 - ploss) * sp0 * sten      # MN — bir sıra

        # ── §5.5.1.2 B kombinasyonu → overstress ve dönme kontrolü ────────
        Fu1 = n_req * sten * (fpeg / 1000.0)         # MN

        Es_MN       = Es_GPa * 1000.0
        delta_sigma = max(0, (Fu1 - Fi)) / (n_req * sten) if n_req * sten > 0 else 0
        delta_L     = delta_sigma * L_ten / Es_MN
        theta_B_rad = delta_L / d if d > 0 else 0
        theta_B_deg = math.degrees(theta_B_rad)

        # ── Blok tepkileri ─────────────────────────────────────────────────
        Ra_A1  = N_A1/2 - M_A1/e + Fi
        Ra_A2  = N_A2/2 - M_A2/e + Fi
        Ra_min = min(Ra_A1, Ra_A2)
        Rb_A   = max(N_A1/2 + M_A1/e + Fi,
                     N_A2/2 + M_A2/e + Fi)
        Rb_B   = max(N_B1/2 + M_B1/e + Fi,
                     N_B2/2 + M_B2/e + Fi)
        Ra_B   = min(N_B1/2 - M_B1/e + Fi,
                     N_B2/2 - M_B2/e + Fi)

        # ── §5.5.2 Blok yüzey alanı ───────────────────────────────────────
        fcf = fc28 + 20.0

        # r1=ba/a0, r2=bb/b0: kucuk/buyuk (SETRA §5.5.2.1)
        r1 = ba / a0 if a0 > 0 else 0
        r2 = bb / b0 if b0 > 0 else 0
        arg_k = (1 - 4*r1/3) * (1 - 4*r2/3)
        raw_K = (3 - (4/3)*(r1 + r2)) * math.sqrt(max(0.0, arg_k))
        K_blk = min(3.3, 1 + raw_K)
        # 1 MPa = 1 MN/m2: fcmax dogrudan MN/m2 biriminde
        fbu_A   = 0.85 * fcf / (1.5 * 1.0)
        fclim_A = K_blk * fbu_A
        fcmax_A = min(fclim_A, 0.75*fc28, fcf)
        S_A     = Rb_A / fcmax_A if fcmax_A > 0 else 999.0
        fbu_B   = 0.85 * fcf / (1.15 * 0.85)
        fclim_B = K_blk * fbu_B
        fcmax_B = min(fclim_B, 0.98*fc28, fcf)
        S_B     = Rb_B / fcmax_B if fcmax_B > 0 else 999.0

        S_prov = 2 * ba * bb

        # ── §5.5.3 Ayak dönme açısı ───────────────────────────────────────
        EI_MN       = 35000.0 * Ip
        theta_A_deg = math.degrees(max(M_A1, M_A2) * hp / EI_MN)

        lf_hp = lf + hp

        ecc_A = max(M_A1/N_A1 if N_A1 > 0 else 0,
                    M_A2/N_A2 if N_A2 > 0 else 0)

        ok_Ra    = Ra_min >= 0
        ok_blk_A = S_prov >= S_A
        ok_blk_B = S_prov >= S_B
        ok_th_A  = theta_A_deg <= 1.0
        ok_th_B  = theta_B_deg <= 1.0
        ok_dyn   = lf_hp <= 180.0
        genel_ok = all([ok_Ra, ok_blk_A, ok_blk_B, ok_th_A, ok_th_B])

        s.update({
            "P": P, "dg": dg, "G_max": G_max, "G_min": G_min,
            "Qprc1max": Qprc1max, "Qprc1min": Qprc1min,
            "Qpra1_kN": Qpra1_kN, "Qpra2": Qpra2,
            "Qw_kN": Qw_kN, "FA": FA,
            "M_A1": M_A1*1000, "N_A1": N_A1*1000,
            "M_A2": M_A2*1000, "N_A2": N_A2*1000,
            "M_B1": M_B1*1000, "N_B1": N_B1*1000,
            "M_B2": M_B2*1000, "N_B2": N_B2*1000,
            "ecc_A": ecc_A, "ecc_limit": e/2,
            "stitch_needed": ecc_A > e/2,
            "sp0": sp0*1000,
            "nxs_A_req_mm2": nxs_A_req*1e6,
            "n_req": n_req,
            "Fi_kN": Fi*1000, "Fu1_kN": Fu1*1000,
            "Ra_min_kN": Ra_min*1000,
            "Ra_A1_kN": Ra_A1*1000, "Ra_A2_kN": Ra_A2*1000,
            "Ra_B_kN": Ra_B*1000,
            "Rb_A_kN": Rb_A*1000, "Rb_B_kN": Rb_B*1000,
            "theta_B_deg": theta_B_deg,
            "fcf": fcf, "K_blk": K_blk,
            "fbu_A": fbu_A, "fcmax_A_MPa": fcmax_A,
            "fbu_B": fbu_B, "fcmax_B_MPa": fcmax_B,
            "S_A_m2": S_A, "S_B_m2": S_B, "S_prov_m2": S_prov,
            "EI_GNm2": EI_MN/1000,
            "theta_A_deg": theta_A_deg,
            "lf_hp": lf_hp,
            "ok_Ra": ok_Ra, "ok_blk_A": ok_blk_A, "ok_blk_B": ok_blk_B,
            "ok_th_A": ok_th_A, "ok_th_B": ok_th_B,
            "ok_dyn": ok_dyn, "genel_ok": genel_ok,
        })
        self.sonuclar = s
        return s


# ═══════════════════════════════════════════════════════════════════════════════
#  RENK PALETİ
# ═══════════════════════════════════════════════════════════════════════════════

R = {
    "bg":      "#F5F4F0",
    "panel":   "#FFFFFF",
    "header":  "#1A1A2E",
    "accent":  "#2D6FA3",
    "ok":      "#27500A",  "ok_bg":   "#EAF3DE",
    "fail":    "#791F1F",  "fail_bg": "#FCEBEB",
    "warn":    "#633806",  "warn_bg": "#FAEEDA",
    "label":   "#555555",
    "mono":    "Courier",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  ANA UYGULAMA
# ═══════════════════════════════════════════════════════════════════════════════

class Uygulama(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("SETRA Bölüm 5 — Dengeli Konsol Stabilite Hesabı")
        self.configure(bg=R["bg"])
        self.resizable(True, True)
        self.minsize(1000, 680)
        self.girdiler: dict[str, tk.StringVar] = {}
        self.sonuclar: dict = {}

        self._stil()
        self._menu()
        self._durum()
        self._duzen()

    # ── Stil ──────────────────────────────────────────────────────────────────
    def _stil(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("TNotebook",     background=R["bg"], borderwidth=0)
        s.configure("TNotebook.Tab", padding=[14, 6], font=("Segoe UI", 9))

    # ── Menü ──────────────────────────────────────────────────────────────────
    def _menu(self):
        mb = tk.Menu(self)
        d  = tk.Menu(mb, tearoff=0)
        d.add_command(label="PDF Rapor Kaydet…",  command=self.pdf_kaydet)
        d.add_command(label="Excel Kaydet…",       command=self.excel_kaydet)
        d.add_separator()
        d.add_command(label="Çıkış",               command=self.destroy)
        mb.add_cascade(label="Dosya", menu=d)
        y  = tk.Menu(mb, tearoff=0)
        y.add_command(label="Hakkında",            command=self._hakkinda)
        mb.add_cascade(label="Yardım", menu=y)
        self.config(menu=mb)

    def _hakkinda(self):
        messagebox.showinfo("Hakkında",
            "SETRA §5 — Dengeli Konsol Stabilite\n"
            "BPEL 91 / BAEL 91 / SETRA GB Chapter 5\n\n"
            "§5.3 Yükler  ·  §5.4 Kombinasyonlar\n"
            "§5.5.1 Tendon  ·  §5.5.2 Blok  ·  §5.5.3 Ayak\n\n"
            "Gerekli: pip install reportlab openpyxl")

    # ── Durum çubuğu ──────────────────────────────────────────────────────────
    def _durum(self):
        self.durum_var = tk.StringVar(
            value="Parametreleri girin ve Hesapla butonuna basın.")
        tk.Label(self, textvariable=self.durum_var, anchor="w",
                 bg=R["header"], fg="#AAAACC",
                 font=("Segoe UI", 9), padx=12, pady=5
                 ).pack(side="bottom", fill="x")

    # ── Ana düzen ─────────────────────────────────────────────────────────────
    def _duzen(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=10, pady=(8, 4))
        self.tab_g = tk.Frame(self.nb, bg=R["bg"])
        self.tab_s = tk.Frame(self.nb, bg=R["bg"])
        self.nb.add(self.tab_g, text="  Parametreler  ")
        self.nb.add(self.tab_s, text="  Sonuçlar  ")
        self._girdi()
        self._sonuc_sekme()

    # ═════════════════════════════════════════════════════════════════════════
    #  GİRDİ SEKMESİ
    # ═════════════════════════════════════════════════════════════════════════

    def _girdi(self):
        f = self.tab_g
        f.columnconfigure((0, 1, 2), weight=1)

        sol = tk.Frame(f, bg=R["bg"])
        sol.grid(row=0, column=0, sticky="nsew", padx=(4,2), pady=4)
        self._p(sol, "Konsol Geometrisi  (§5.3.1)", [
            ("Yarım konsol uzunluğu  l_f",   "lf",    "40.0",  "m"),
            ("Kesit alanı (SOP)  B₁",        "B1",    "9.50",  "m²"),
            ("Taç kesit alanı  B₀",          "B0",    "5.20",  "m²"),
            ("Birim ağırlık  γ",             "gamma", "24.5",  "kN/m³"),
            ("Üst başlık genişliği  b",      "bslab", "11.0",  "m"),
        ], "Krawsky: P = (B₁+2B₀)·γ·l_f / 3")

        self._p(sol, "Geçici Destek  (§5.5)", [
            ("Blok sırası arası  e",          "e_blk", "3.20", "m"),
            ("Tendon – karşı blok  d",        "d_blk", "2.00", "m"),
            ("Kayıp oranı  p",                "p_loss","20",   "%"),
            ("Tendon  f_prg  (çekme)",        "fprg",  "1860", "MPa"),
            ("Tendon  f_peg  (akma)",         "fpeg",  "1600", "MPa"),
            ("Tendon kesit alanı  s",         "s_ten", "1500", "mm²"),
            ("Tendon serbest boyu  L",        "L_ten", "12.0", "m"),
            ("Tendon elastisite  E_s",        "Es",    "195",  "GPa"),
        ], "σ_p0 = min(0.8·f_prg , 0.9·f_peg)  ·  min. 2 çift tendon/sıra")

        ort = tk.Frame(f, bg=R["bg"])
        ort.grid(row=0, column=1, sticky="nsew", padx=2, pady=4)
        self._p(ort, "Bilinen İnşaat Yükleri  (§5.3.2.1)", [
            ("Kalıp arabası  Q_PRC1  (nominal)", "Qprc1", "800", "kN"),
            ("Kalıp arabası konum  d_ex",         "dex",   "2.0", "m"),
        ], "Q_PRC1max = +%6  ·  Q_PRC1min = −%4")

        self._p(ort, "Rastgele İnşaat Yükleri  (§5.3.2.2)", [
            ("Yayılı yük  Q_PRA1",              "Qpra1", "0.20", "kN/m²"),
            ("Konsantrik  Q_PRA2  (0 = oto)",   "Qpra2", "0",    "kN"),
            ("Ek rüzgar  Q_w  (L > 120 m)",     "Qw",    "0.0",  "kN/m²"),
        ], "Q_PRA2 oto = 50 + 5·b  [kN]")

        self._p(ort, "Kazara Yük  (§5.3.3)", [],
                "F_A = 2·Q_PRC1max  (dinamik katsayı ×2)\n"
                "B kombinasyonunda, yön ters olarak uygulanır.")

        self._p(ort, "Ayak Geometrisi  (§5.5.3)", [
            ("Ayak yüksekliği  h_p",   "hp",   "12.0", "m"),
            ("Atalet momenti  I_p",    "Ip",   "3.50", "m⁴"),
            ("Beton dayanımı  f_c28",  "fc28", "35",   "MPa"),
        ], "θ = M·h_p/(E_c·I_p) ≤ 1°  ·  E_c = 35 000 MPa")

        sag = tk.Frame(f, bg=R["bg"])
        sag.grid(row=0, column=2, sticky="nsew", padx=(2,4), pady=4)
        self._p(sag, "Geçici Blok Boyutları  (§5.5.2)", [
            ("Blok boyutu  a  (uzunluk)",  "blk_a", "0.90", "m"),
            ("Blok boyutu  b  (genişlik)", "blk_b", "0.90", "m"),
            ("Destek yüzeyi  a₀",          "a0",    "1.50", "m"),
            ("Destek yüzeyi  b₀",          "b0",    "1.50", "m"),
        ], "S_sağlanan = 2·a·b  ≥  R_b / f_cmax  (BAEL 91 A.8.4)")

        # Butonlar
        btn = tk.Frame(f, bg=R["bg"])
        btn.grid(row=1, column=0, columnspan=3, pady=(4, 10))
        for txt, cmd, clr in [
            ("  HESAPLA  →",    self.hesapla,      R["header"]),
            ("  PDF Rapor",     self.pdf_kaydet,   R["accent"]),
            ("  Excel Çıktı",   self.excel_kaydet, "#217346"),
        ]:
            tk.Button(btn, text=txt, command=cmd,
                      bg=clr, fg="white",
                      font=("Segoe UI", 10, "bold"),
                      relief="flat", padx=18, pady=10,
                      cursor="hand2").pack(side="left", padx=6)

    def _p(self, parent, baslik, alanlar, not_str=""):
        frame = tk.LabelFrame(parent, text=f"  {baslik}  ",
                              bg=R["panel"], fg=R["header"],
                              font=("Segoe UI", 9, "bold"),
                              relief="groove", bd=1, padx=10, pady=8)
        frame.pack(fill="x", pady=(0, 6))
        for etiket, anahtar, varsayilan, birim in alanlar:
            row = tk.Frame(frame, bg=R["panel"])
            row.pack(fill="x", pady=2)
            tk.Label(row, text=etiket, bg=R["panel"], fg=R["label"],
                     font=("Segoe UI", 9), width=28, anchor="w").pack(side="left")
            var = tk.StringVar(value=varsayilan)
            self.girdiler[anahtar] = var
            tk.Entry(row, textvariable=var, width=9,
                     font=(R["mono"], 9), relief="solid", bd=1
                     ).pack(side="left", padx=(4, 2))
            tk.Label(row, text=birim, bg=R["panel"],
                     fg="#999", font=("Segoe UI", 8)).pack(side="left")
        if not_str:
            tk.Label(frame, text=not_str, bg=R["panel"],
                     fg="#888", font=("Segoe UI", 8, "italic"),
                     anchor="w", justify="left").pack(fill="x", pady=(4, 0))

    # ═════════════════════════════════════════════════════════════════════════
    #  SONUÇ SEKMESİ
    # ═════════════════════════════════════════════════════════════════════════

    def _sonuc_sekme(self):
        f  = self.tab_s
        cv = tk.Canvas(f, bg=R["bg"], highlightthickness=0)
        sb = ttk.Scrollbar(f, orient="vertical", command=cv.yview)
        cv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True)
        self.ic = tk.Frame(cv, bg=R["bg"])
        self._ic_id = cv.create_window((0, 0), window=self.ic, anchor="nw")

        def _r(e=None):
            cv.configure(scrollregion=cv.bbox("all"))
            cv.itemconfig(self._ic_id, width=cv.winfo_width())

        self.ic.bind("<Configure>", _r)
        cv.bind("<Configure>", _r)
        cv.bind_all("<MouseWheel>",
                    lambda e: cv.yview_scroll(-1*(e.delta//120), "units"))

    def _temizle(self):
        for w in self.ic.winfo_children():
            w.destroy()

    def _goster(self, s):
        ic = self.ic
        ic.columnconfigure((0, 1, 2), weight=1)

        # Banner
        ok  = s["genel_ok"]
        msg = ("✓  TÜM KONTROLLER SAĞLANDI" if ok
               else "✗  BİR VEYA DAHA FAZLA KONTROL BAŞARISIZ")
        b = tk.Frame(ic, bg=R["ok_bg"] if ok else R["fail_bg"], pady=12)
        b.grid(row=0, column=0, columnspan=3, sticky="ew", padx=8, pady=(8,4))
        tk.Label(b, text=msg, bg=b.cget("bg"),
                 fg=R["ok"] if ok else R["fail"],
                 font=("Segoe UI", 12, "bold")).pack()

        if not s["ok_dyn"]:
            w = tk.Frame(ic, bg=R["warn_bg"], pady=6)
            w.grid(row=1, column=0, columnspan=3,
                   sticky="ew", padx=8, pady=(0, 4))
            tk.Label(w,
                     text=f"⚠  l_f + h_p = {s['lf_hp']:.0f} m > 180 m  →  "
                          "Türbülanslı rüzgar dinamik analizi zorunludur (§5.4.3)",
                     bg=R["warn_bg"], fg=R["warn"],
                     font=("Segoe UI", 9, "bold")).pack()

        def kutu(row, col, baslik):
            f = tk.LabelFrame(ic, text=f"  {baslik}  ",
                              bg=R["panel"], fg=R["header"],
                              font=("Segoe UI", 9, "bold"),
                              relief="groove", bd=1, padx=10, pady=8)
            f.grid(row=row, column=col, sticky="nsew", padx=6, pady=4)
            return f

        def veri(parent, etiket, deger, durum=None):
            bg = R["panel"]
            if durum == "ok":   bg = R["ok_bg"]
            elif durum == "fail": bg = R["fail_bg"]
            elif durum == "warn": bg = R["warn_bg"]
            row = tk.Frame(parent, bg=bg)
            row.pack(fill="x", pady=1)
            tk.Label(row, text=etiket, bg=bg, fg=R["label"],
                     font=("Segoe UI", 9), anchor="w",
                     width=33).pack(side="left")
            tk.Label(row, text=str(deger), bg=bg, fg="#111",
                     font=(R["mono"], 9, "bold"),
                     anchor="e").pack(side="right")

        # ── Satır 2 ──
        k = kutu(2, 0, "§5.3 — Yük Büyüklükleri  (Krawsky)")
        veri(k, "Yarım konsol ağırlığı P",    f"{s['P']:.1f} kN")
        veri(k, "Ağırlık merkezi d",           f"{s['dg']:.3f} m")
        veri(k, "G_max (+%2)",                 f"{s['G_max']:.1f} kN")
        veri(k, "G_min (−%2)",                 f"{s['G_min']:.1f} kN")
        veri(k, "Q_PRC1max (+%6)",             f"{s['Qprc1max']:.1f} kN")
        veri(k, "Q_PRC1min (−%4)",             f"{s['Qprc1min']:.1f} kN")
        veri(k, "Q_PRA1 (toplam yayılı)",      f"{s['Qpra1_kN']:.1f} kN")
        veri(k, "Q_PRA2 = 50+5b",             f"{s['Qpra2']:.1f} kN")
        veri(k, "Q_w (ek rüzgar, toplam)",     f"{s['Qw_kN']:.1f} kN")
        veri(k, "F_A = 2×Q_PRC1max",          f"{s['FA']:.1f} kN")

        m = kutu(2, 1, "§5.4 — Kombinasyon M ve N")
        veri(m, "M_A1  [Geçici]",  f"{s['M_A1']:.1f} kN·m")
        veri(m, "N_A1",             f"{s['N_A1']:.1f} kN")
        veri(m, "M_A2  [Geçici]",  f"{s['M_A2']:.1f} kN·m")
        veri(m, "N_A2",             f"{s['N_A2']:.1f} kN")
        veri(m, "M_B1  [Kazara]",  f"{s['M_B1']:.1f} kN·m")
        veri(m, "N_B1",             f"{s['N_B1']:.1f} kN")
        veri(m, "M_B2  [Kazara]",  f"{s['M_B2']:.1f} kN·m")
        veri(m, "N_B2",             f"{s['N_B2']:.1f} kN")
        dur = "warn" if s["stitch_needed"] else "ok"
        sim = ">" if s["stitch_needed"] else "<"
        veri(m, "Eksantrisite M/N  (A, maks)",
             f"{s['ecc_A']:.3f} m  {sim}  {s['ecc_limit']:.2f} m (e/2)", dur)

        t = kutu(2, 2, "§5.5.1 — Tendon Boyutlandırması")
        veri(t, "σ_p0 = min(0.8f_prg, 0.9f_peg)",  f"{s['sp0']:.1f} MPa")
        veri(t, "n×s gerekli  (A komb.)",            f"{s['nxs_A_req_mm2']:.0f} mm²")
        veri(t, "Gerekli tendon sayısı n  /sıra",    f"{s['n_req']} adet")
        veri(t, "F_i  (bir sıra kuvveti)",            f"{s['Fi_kN']:.1f} kN")
        veri(t, "F_u1  (limit, γ_p=1.0)",            f"{s['Fu1_kN']:.1f} kN")
        dur_ra = "ok" if s["ok_Ra"] else "fail"
        veri(t, "R_A minimum  (A komb.)",
             f"{s['Ra_min_kN']:.1f} kN  "
             f"{'≥ 0  ✓' if s['ok_Ra'] else '< 0  ✗'}", dur_ra)
        veri(t, "R_A  (B komb., bilgi)",
             f"{s['Ra_B_kN']:.0f} kN  "
             f"({'kalkma — normal' if s['Ra_B_kN']<0 else 'kalkma yok'})")
        dur_tb = "ok" if s["ok_th_B"] else "warn"
        veri(t, "Dönme açısı θ_B  (B komb.)",
             f"{s['theta_B_deg']:.4f}°  "
             f"{'≤ 1°  ✓' if s['ok_th_B'] else '> 1°  ✗'}", dur_tb)

        # ── Satır 3 ──
        bl = kutu(3, 0, "§5.5.2 — Blok Yüzey Alanı  (BAEL 91 A.8.4)")
        veri(bl, "f_cf = f_c28+20  (hooplama)",    f"{s['fcf']:.0f} MPa")
        veri(bl, "K  yayılma katsayısı  ≤ 3.3",    f"{s['K_blk']:.3f}")
        veri(bl, "f_bu  (A,  γ_b=1.5, θ=1.0)",    f"{s['fbu_A']:.2f} MPa")
        veri(bl, "f_cmax  (A kombinasyonu)",         f"{s['fcmax_A_MPa']:.2f} MPa")
        veri(bl, "Gerekli alan S_A",                 f"{s['S_A_m2']:.5f} m²")
        dur_bA = "ok" if s["ok_blk_A"] else "fail"
        veri(bl, "Sağlanan 2·a·b  vs  S_A",
             f"{s['S_prov_m2']:.4f}  "
             f"{'≥' if s['ok_blk_A'] else '<'}  {s['S_A_m2']:.5f} m²", dur_bA)
        veri(bl, "f_bu  (B,  γ_b=1.15, θ=0.85)",  f"{s['fbu_B']:.2f} MPa")
        veri(bl, "f_cmax  (B kombinasyonu)",         f"{s['fcmax_B_MPa']:.2f} MPa")
        veri(bl, "Gerekli alan S_B",                 f"{s['S_B_m2']:.5f} m²")
        dur_bB = "ok" if s["ok_blk_B"] else "fail"
        veri(bl, "Sağlanan 2·a·b  vs  S_B",
             f"{s['S_prov_m2']:.4f}  "
             f"{'≥' if s['ok_blk_B'] else '<'}  {s['S_B_m2']:.5f} m²", dur_bB)

        ay = kutu(3, 1, "§5.5.3 — Ayak Dönme Açısı")
        veri(ay, "E_c × I_p",  f"{s['EI_GNm2']:.2f} GN·m²")
        dur_tA = "ok" if s["ok_th_A"] else "warn"
        veri(ay, "θ_A  (A komb.)  ≤ 1°",
             f"{s['theta_A_deg']:.4f}°  "
             f"{'✓' if s['ok_th_A'] else '✗'}", dur_tA)
        dur_dyn = "ok" if s["ok_dyn"] else "warn"
        veri(ay, "l_f + h_p  (>180 m → dinamik)",
             f"{s['lf_hp']:.1f} m  →  "
             f"{'Statik yeterli' if s['ok_dyn'] else 'DİNAMİK ANALİZ!'}", dur_dyn)
        tk.Label(ay,
                 text="Tip A → temel kombinasyon (γ_b=1.5, γ_s=1.15)\n"
                      "Tip B → kazara kombinasyon (γ_b=1.15, γ_s=1.00)\n"
                      "İnşaat fazı çatlak sınırı: σ_s ≤ 400 MPa",
                 bg=R["panel"], fg="#888",
                 font=("Segoe UI", 8, "italic"),
                 justify="left", anchor="w").pack(fill="x", pady=(6, 0))

        oz = kutu(3, 2, "Özet Kontrol Tablosu")
        for ad, ok_flag in [
            ("Blok A ayrılmaması  R_A ≥ 0  (A komb.)", s["ok_Ra"]),
            ("Blok yüzey alanı — A kombinasyonu",        s["ok_blk_A"]),
            ("Blok yüzey alanı — B kombinasyonu",        s["ok_blk_B"]),
            ("Ayak dönme açısı θ_A ≤ 1°  (A komb.)",   s["ok_th_A"]),
            ("Tendon dönme açısı θ_B ≤ 1°  (B komb.)", s["ok_th_B"]),
        ]:
            veri(oz, ad,
                 "TAMAM  ✓" if ok_flag else "BAŞARISIZ  ✗",
                 "ok" if ok_flag else "fail")
        if not s["ok_dyn"]:
            veri(oz, "Dinamik analiz  (§5.4.3)",
                 f"ZORUNLU  ({s['lf_hp']:.0f} m > 180 m)", "warn")

    # ═════════════════════════════════════════════════════════════════════════
    #  HESAPLA
    # ═════════════════════════════════════════════════════════════════════════

    def _oku(self) -> dict:
        anahtarlar = [
            "lf","B1","B0","gamma","bslab",
            "e_blk","d_blk","p_loss","fprg","fpeg","s_ten","L_ten","Es",
            "hp","Ip","fc28",
            "blk_a","blk_b","a0","b0",
            "Qprc1","dex","Qpra1","Qpra2","Qw",
        ]
        p = {}
        for k in anahtarlar:
            try:
                p[k] = float(self.girdiler[k].get())
            except ValueError:
                raise ValueError(f"'{k}' alanında geçersiz değer!")
        return p

    def hesapla(self):
        try:
            p = self._oku()
        except ValueError as e:
            messagebox.showerror("Giriş Hatası", str(e))
            return
        try:
            s = SetraHesap(p).hesapla()
            self.sonuclar = s
            self._temizle()
            self._goster(s)
            self.nb.select(self.tab_s)
            self.durum_var.set(
                "✓  Hesap tamamlandı — tüm kontroller sağlandı."
                if s["genel_ok"] else
                "✗  Hesap tamamlandı — bir veya daha fazla kontrol başarısız!")
        except Exception as e:
            messagebox.showerror("Hesap Hatası", str(e))

    # ═════════════════════════════════════════════════════════════════════════
    #  PDF
    # ═════════════════════════════════════════════════════════════════════════

    def pdf_kaydet(self):
        if not self.sonuclar:
            messagebox.showwarning("Uyarı", "Önce hesap yapın.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile="setra_b5_stabilite.pdf")
        if not path:
            return
        try:
            self._pdf(path)
            messagebox.showinfo("PDF Kaydedildi", path)
        except Exception as e:
            messagebox.showerror("PDF Hatası", str(e))

    def _pdf(self, path):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph,
                                         Spacer, Table, TableStyle,
                                         HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        s   = self.sonuclar
        doc = SimpleDocTemplate(path, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()

        NAVY = colors.HexColor("#1A1A2E")
        BLUE = colors.HexColor("#2D6FA3")
        LGRY = colors.HexColor("#F5F5F5")

        H1 = ParagraphStyle("h1", parent=styles["Heading1"],
                            textColor=NAVY, fontSize=13, spaceAfter=4)
        H2 = ParagraphStyle("h2", parent=styles["Heading2"],
                            textColor=BLUE, fontSize=10,
                            spaceBefore=8, spaceAfter=3)
        SM = ParagraphStyle("sm", parent=styles["Normal"],
                            fontSize=8, textColor=colors.gray, leading=11)

        def tbl(heads, rows, widths=None):
            data = [heads] + rows
            t = Table(data, colWidths=widths)
            ln = colors.HexColor("#CCCCCC")
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(-1,0), NAVY),
                ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
                ("FONTNAME",      (0,0),(-1,-1),"Helvetica"),
                ("FONTSIZE",      (0,0),(-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LGRY]),
                ("GRID",          (0,0),(-1,-1), 0.3, ln),
                ("VALIGN",        (0,0),(-1,-1),"MIDDLE"),
                ("TOPPADDING",    (0,0),(-1,-1), 4),
                ("BOTTOMPADDING", (0,0),(-1,-1), 4),
            ]))
            return t

        cw3 = [5.5*cm, 6*cm, 4*cm]
        akis = []

        akis.append(Paragraph(
            "SETRA Bölüm 5 — Dengeli Konsol Stabilite Hesabı", H1))
        akis.append(Paragraph(
            f"SETRA GB Chapter 5  ·  BPEL 91  ·  BAEL 91  |  "
            f"{datetime.date.today().strftime('%d.%m.%Y')}  "
            f"{datetime.datetime.now().strftime('%H:%M')}", SM))
        akis.append(HRFlowable(width="100%", thickness=1.5,
                                color=NAVY, spaceAfter=8))

        ok   = s["genel_ok"]
        renk = colors.HexColor("#27500A" if ok else "#791F1F")
        akis.append(Paragraph(
            f"<b>Genel Sonuç:  "
            f"{'TÜM KONTROLLER SAĞLANDI  ✓' if ok else 'BİR VEYA DAHA FAZLA KONTROL BAŞARISIZ  ✗'}"
            f"</b>",
            ParagraphStyle("gs", parent=styles["Normal"],
                           textColor=renk, fontSize=11)))

        if not s["ok_dyn"]:
            akis.append(Paragraph(
                f"⚠  l_f + h_p = {s['lf_hp']:.0f} m > 180 m  →  "
                "Dinamik analiz zorunlu (§5.4.3)",
                ParagraphStyle("w", parent=styles["Normal"],
                               textColor=colors.HexColor("#633806"),
                               fontSize=9)))
        akis.append(Spacer(1, 0.3*cm))

        akis.append(Paragraph("§5.3.1  Krawsky Formülleri", H2))
        akis.append(tbl(
            ["Büyüklük", "Formül", "Değer"],
            [["P", "(B₁+2B₀)·γ·l_f/3",           f"{s['P']:.1f} kN"],
             ["d", "(B₁+5B₀)·l_f/[4(B₁+2B₀)]",  f"{s['dg']:.3f} m"],
             ["G_max (+%2)", "1.02·P",             f"{s['G_max']:.1f} kN"],
             ["G_min (−%2)", "0.98·P",             f"{s['G_min']:.1f} kN"],
             ["Q_PRC1max (+%6)", "1.06·Q_PRC1",   f"{s['Qprc1max']:.1f} kN"],
             ["Q_PRC1min (−%4)", "0.96·Q_PRC1",   f"{s['Qprc1min']:.1f} kN"],
             ["Q_PRA2", "50+5·b",                  f"{s['Qpra2']:.1f} kN"],
             ["F_A (kazara)", "2·Q_PRC1max",       f"{s['FA']:.1f} kN"]],
            cw3))

        akis.append(Paragraph("§5.4  Kombinasyon M ve N Değerleri", H2))
        akis.append(tbl(
            ["Komb.", "Tip", "M  [kN·m]", "N  [kN]", "M/N  [m]"],
            [["A1","Geçici",f"{s['M_A1']:.0f}",f"{s['N_A1']:.0f}",
              f"{s['M_A1']/s['N_A1']:.3f}"],
             ["A2","Geçici",f"{s['M_A2']:.0f}",f"{s['N_A2']:.0f}",
              f"{s['M_A2']/s['N_A2']:.3f}"],
             ["B1","Kazara",f"{s['M_B1']:.0f}",f"{s['N_B1']:.0f}",
              f"{s['M_B1']/s['N_B1']:.3f}"],
             ["B2","Kazara",f"{s['M_B2']:.0f}",f"{s['N_B2']:.0f}",
              f"{s['M_B2']/s['N_B2']:.3f}"]],
            [1.5*cm, 4*cm, 3.5*cm, 3.5*cm, 3*cm]))

        akis.append(Paragraph("§5.5.1  Tendon Boyutlandırması", H2))
        akis.append(tbl(
            ["Büyüklük", "Açıklama", "Değer"],
            [["σ_p0","min(0.8f_prg, 0.9f_peg)",        f"{s['sp0']:.1f} MPa"],
             ["n×s req. (A)","A kombinasyonu",           f"{s['nxs_A_req_mm2']:.0f} mm²"],
             ["n  /sıra","Çift sayı, min 2",             f"{s['n_req']} adet"],
             ["F_i","n·(1-p)·σ_p0·s",                  f"{s['Fi_kN']:.1f} kN"],
             ["F_u1","n·s·f_peg/γ_p (γ_p=1.0)",        f"{s['Fu1_kN']:.1f} kN"],
             ["R_A min","N/2−M/e+F_i",
              f"{s['Ra_min_kN']:.1f} kN  {'TAMAM ✓' if s['ok_Ra'] else 'BAŞARISIZ ✗'}"],
             ["θ_B  (B komb.)","Overtension dönme",
              f"{s['theta_B_deg']:.4f}°  {'≤1° ✓' if s['ok_th_B'] else '>1° ✗'}"]],
            cw3))

        akis.append(Paragraph("§5.5.2  Blok Yüzey Alanı  (BAEL 91 A.8.4)", H2))
        akis.append(tbl(
            ["Büyüklük", "Değer", "Kontrol"],
            [["f_cf = f_c28+20",     f"{s['fcf']:.0f} MPa",        ""],
             ["K yayılma  ≤3.3",     f"{s['K_blk']:.3f}",          ""],
             ["f_cmax  (A)",          f"{s['fcmax_A_MPa']:.2f} MPa",""],
             ["S_A  [m²]",            f"{s['S_A_m2']:.5f}",
              "TAMAM ✓" if s["ok_blk_A"] else "YETERSİZ ✗"],
             ["f_cmax  (B)",          f"{s['fcmax_B_MPa']:.2f} MPa",""],
             ["S_B  [m²]",            f"{s['S_B_m2']:.5f}",
              "TAMAM ✓" if s["ok_blk_B"] else "YETERSİZ ✗"],
             ["2·a·b  [m²]",          f"{s['S_prov_m2']:.4f}",      ""]],
            [5.5*cm, 4.5*cm, 4*cm]))

        akis.append(Paragraph("§5.5.3  Ayak Dönme Açısı", H2))
        akis.append(tbl(
            ["Büyüklük", "Değer", "Kontrol"],
            [["EI_p", f"{s['EI_GNm2']:.2f} GN·m²", ""],
             ["θ_A  ≤ 1°", f"{s['theta_A_deg']:.4f}°",
              "TAMAM ✓" if s["ok_th_A"] else "KONTROL ET ✗"],
             ["l_f + h_p", f"{s['lf_hp']:.1f} m",
              "Statik yeterli" if s["ok_dyn"] else "DİNAMİK ANALİZ zorunlu"]],
            cw3))

        akis.append(HRFlowable(width="100%", thickness=0.5,
                                color=colors.HexColor("#CCCCCC"), spaceAfter=6))
        akis.append(Paragraph("Özet Kontrol Tablosu", H2))
        akis.append(tbl(
            ["Kontrol", "Sonuç"],
            [[ad, "TAMAM ✓" if ok else "BAŞARISIZ ✗"] for ad, ok in [
                ("Blok A ayrılmaması  R_A ≥ 0  (A komb.)", s["ok_Ra"]),
                ("Blok yüzey alanı — A kombinasyonu",       s["ok_blk_A"]),
                ("Blok yüzey alanı — B kombinasyonu",       s["ok_blk_B"]),
                ("Ayak θ_A ≤ 1°  (A komb.)",               s["ok_th_A"]),
                ("Tendon θ_B ≤ 1°  (B komb.)",             s["ok_th_B"]),
            ]],
            [12*cm, 3.5*cm]))

        akis.append(Spacer(1, 0.4*cm))
        akis.append(Paragraph(
            "Bu hesap SETRA GB Chapter 5, BPEL 91 ve BAEL 91 esaslarına göre hazırlanmıştır. "
            "Giriş parametrelerinin ve yük kombinasyonlarının proje özelinde "
            "bağımsız olarak doğrulanması mühendislik sorumluluğundadır.", SM))

        doc.build(akis)

    # ═════════════════════════════════════════════════════════════════════════
    #  EXCEL
    # ═════════════════════════════════════════════════════════════════════════

    def excel_kaydet(self):
        if not self.sonuclar:
            messagebox.showwarning("Uyarı", "Önce hesap yapın.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="setra_b5_stabilite.xlsx")
        if not path:
            return
        try:
            self._excel(path)
            messagebox.showinfo("Excel Kaydedildi", path)
        except Exception as e:
            messagebox.showerror("Excel Hatası", str(e))

    def _excel(self, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "SETRA §5"
        s = self.sonuclar

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14

        thin = Side(style="thin", color="CCCCCC")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

        r = [0]

        def C(row, col, val, bold=False, bg=None, fg="111111",
              sz=9, align="left"):
            c = ws.cell(row, col, val)
            c.font      = Font(bold=bold, size=sz, color=fg, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=bg) if bg else PatternFill()
            c.alignment = Alignment(horizontal=align, vertical="center",
                                     wrap_text=True)
            c.border    = brd
            return c

        def baslik(txt, bg="1A1A2E", fg="FFFFFF"):
            r[0] += 1
            ws.merge_cells(f"A{r[0]}:C{r[0]}")
            C(r[0], 1, txt, True, bg, fg, 10)
            r[0] += 1

        def veri(lbl, val, durum=None):
            bg = {"ok":"EAF3DE","fail":"FCEBEB","warn":"FAEEDA"}.get(
                durum, "FFFFFF" if r[0]%2==1 else "F5F5F5")
            C(r[0], 1, lbl, bg=bg)
            C(r[0], 2, val, True, bg, align="right")
            r[0] += 1

        def bos(): r[0] += 1

        baslik("SETRA Bölüm 5 — Dengeli Konsol Stabilite Hesabı")
        veri("Tarih",      datetime.date.today().strftime("%d.%m.%Y"))
        veri("Referans",   "SETRA GB Ch.5 / BPEL 91 / BAEL 91")
        veri("Genel Sonuç",
             "TÜM KONTROLLER SAĞLANDI" if s["genel_ok"] else "BAŞARISIZ",
             "ok" if s["genel_ok"] else "fail")
        bos()

        baslik("§5.3.1 Krawsky Formülleri", "2D6FA3")
        for k,v in [("P  [kN]",f"{s['P']:.1f}"),("d  [m]",f"{s['dg']:.3f}"),
                    ("G_max [kN]",f"{s['G_max']:.1f}"),("G_min [kN]",f"{s['G_min']:.1f}"),
                    ("Q_PRC1max [kN]",f"{s['Qprc1max']:.1f}"),
                    ("Q_PRC1min [kN]",f"{s['Qprc1min']:.1f}"),
                    ("Q_PRA2 [kN]",f"{s['Qpra2']:.1f}"),
                    ("F_A [kN]",f"{s['FA']:.1f}")]:
            veri(k,v)
        bos()

        baslik("§5.4 Kombinasyon M ve N", "2D6FA3")
        for k,v in [("M_A1 [kN·m]",f"{s['M_A1']:.0f}"),("N_A1 [kN]",f"{s['N_A1']:.0f}"),
                    ("M_A2 [kN·m]",f"{s['M_A2']:.0f}"),("N_A2 [kN]",f"{s['N_A2']:.0f}"),
                    ("M_B1 [kN·m]",f"{s['M_B1']:.0f}"),("N_B1 [kN]",f"{s['N_B1']:.0f}"),
                    ("M_B2 [kN·m]",f"{s['M_B2']:.0f}"),("N_B2 [kN]",f"{s['N_B2']:.0f}")]:
            veri(k,v)
        bos()

        baslik("§5.5.1 Tendon Boyutlandırması", "2D6FA3")
        veri("σ_p0 [MPa]",           f"{s['sp0']:.1f}")
        veri("n×s gerekli (A) [mm²]",f"{s['nxs_A_req_mm2']:.0f}")
        veri("n tendon/sıra [adet]",  f"{s['n_req']}")
        veri("F_i [kN]",              f"{s['Fi_kN']:.1f}")
        veri("F_u1 [kN]",             f"{s['Fu1_kN']:.1f}")
        veri("R_A minimum [kN]",
             f"{s['Ra_min_kN']:.1f}  {'TAMAM ✓' if s['ok_Ra'] else 'BAŞARISIZ ✗'}",
             "ok" if s["ok_Ra"] else "fail")
        veri("θ_B (B komb.) [°]",
             f"{s['theta_B_deg']:.4f}  {'≤1° ✓' if s['ok_th_B'] else '>1° ✗'}",
             "ok" if s["ok_th_B"] else "warn")
        bos()

        baslik("§5.5.2 Blok Yüzey Alanı", "2D6FA3")
        veri("f_cf [MPa]",      f"{s['fcf']:.0f}")
        veri("K katsayısı",     f"{s['K_blk']:.3f}")
        veri("f_cmax A [MPa]",  f"{s['fcmax_A_MPa']:.2f}")
        veri("S_A [m²]",        f"{s['S_A_m2']:.5f}")
        veri("f_cmax B [MPa]",  f"{s['fcmax_B_MPa']:.2f}")
        veri("S_B [m²]",        f"{s['S_B_m2']:.5f}")
        veri("2·a·b [m²]",      f"{s['S_prov_m2']:.4f}")
        veri("Blok A komb.",
             "TAMAM ✓" if s["ok_blk_A"] else "YETERSİZ ✗",
             "ok" if s["ok_blk_A"] else "fail")
        veri("Blok B komb.",
             "TAMAM ✓" if s["ok_blk_B"] else "YETERSİZ ✗",
             "ok" if s["ok_blk_B"] else "fail")
        bos()

        baslik("§5.5.3 Ayak Dönme Açısı", "2D6FA3")
        veri("EI_p [GN·m²]",   f"{s['EI_GNm2']:.2f}")
        veri("θ_A [°]",
             f"{s['theta_A_deg']:.4f}  {'≤1° ✓' if s['ok_th_A'] else '>1° ✗'}",
             "ok" if s["ok_th_A"] else "warn")
        veri("l_f + h_p [m]",
             f"{s['lf_hp']:.1f}  {'Statik yeterli' if s['ok_dyn'] else 'DİNAMİK ANALİZ!'}",
             "ok" if s["ok_dyn"] else "warn")
        bos()

        baslik("Özet Kontrol Tablosu")
        for ad, ok in [
            ("Blok A ayrılmaması R_A ≥ 0 (A komb.)", s["ok_Ra"]),
            ("Blok yüzey alanı — A kombinasyonu",      s["ok_blk_A"]),
            ("Blok yüzey alanı — B kombinasyonu",      s["ok_blk_B"]),
            ("Ayak θ_A ≤ 1° (A komb.)",               s["ok_th_A"]),
            ("Tendon θ_B ≤ 1° (B komb.)",             s["ok_th_B"]),
        ]:
            veri(ad, "TAMAM ✓" if ok else "BAŞARISIZ ✗",
                 "ok" if ok else "fail")

        wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = Uygulama()
    app.mainloop()
